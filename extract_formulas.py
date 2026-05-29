import openpyxl

def extract_formulas(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=False)
    sheet = wb['Planets']
    
    print(f"Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
    
    for row in range(1, sheet.max_row + 1):
        row_data = []
        has_data = False
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value is not None:
                has_data = True
                row_data.append(f"{cell.coordinate}:{cell.value}")
        if has_data:
            print(" | ".join(row_data))

if __name__ == "__main__":
    extract_formulas('PlanetFinder.xlsm')
